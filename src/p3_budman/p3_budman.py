# ---------------------------------------------------------------------------- +
#region p3_budman.py The Budget Manager Application.
""" Budget Manager (BudMan) - an app to help users track a financial budget.

    This module is the main entry point for the BudMan application.
"""
#endregion p3_budman.py The Budget Manager Application.
# ---------------------------------------------------------------------------- +
#region Imports
# python standard libraries packages and modules 
import atexit, pathlib, logging, inspect, logging.config  #, logging.handlers

# third-party  packages and module libraries
from config import settings
from openpyxl import Workbook
import p3logging as p3l, p3_utils as p3u

# local packages and module libraries
import budman_view_model.budman_cli_view_datacontext as p3bm_view_dc
import budman_view_model.budman_view_model_interface as p3bmvm
import budman_cli_view.budman_cli_view as p3bmv
logger = logging.getLogger(settings.app_name)
logger.propagate = True
#endregion Imports
# ---------------------------------------------------------------------------- +
#region configure_logging() function
def configure_logging(logger_name : str = settings.app_name, logtest : bool = False) -> None:
    """Setup the application logger."""
    try:
        # Configure logging
        log_config_file = "budget_model_logging_config.jsonc"
        _ = p3l.setup_logging(
            logger_name = logger_name,
            config_file = log_config_file
            )
        p3l.set_log_flag(p3l.LOG_FLAG_PRINT_CONFIG_ERRORS, True)
        logger = logging.getLogger(settings.app_name)
        logger.setLevel(logging.DEBUG)
        logger.info("+ ----------------------------------------------------- +")
        logger.info(f"+ Running {settings.app_name}...")
        logger.info("+ ----------------------------------------------------- +")
        if(logtest): 
            p3l.quick_logging_test(settings.app_name, log_config_file, reload = False)
    except Exception as e:
        logger.error(p3u.exc_msg(configure_logging, e))
        raise
#endregion configure_logging() function
# ---------------------------------------------------------------------------- +
#region log_workbook_info() function
def log_workbook_info(file_name : str = "unknown",wb : Workbook = None) -> None:  
    if wb is None or not isinstance(wb, Workbook):
        logger.error(f"Workbook({file_name}): None or not a Workbook.")
        return
    logger.info(f"Workbook({file_name}): with sheets: {str(wb.sheetnames)}")
    # sheet = wb.active
    # r = sheet.max_row
    # c = sheet.max_column
    logger.info(f"  Active sheet({wb.active.title}): " 
                f"({wb.active.max_row} x {wb.active.max_column})")
#endregion log_workbook_info() function
# ---------------------------------------------------------------------------- +
#region budman_app_cli_cmdloop() function
def budman_app_cli_cmdloop(startup : bool = True) -> None:
    """CLI cmdloop function."""
    try:
        # create and initialize view model
        bmvm = p3bmvm.BudgetManagerViewModelInterface()
        bmvm.initialize(load_user_store=True) # Initialize the BudgetModelCommandViewModel
        # create and initialize a data context, for the view model
        bm_cliview_dc = p3bm_view_dc.BudgetManagerCLIViewDataContext(bmvm)
        bm_cliview_dc.initialize(cp=bmvm.BMVM_execute_cmd,dc=bmvm.data_context) 
        # create and initialize the view
        clview = p3bmv.BudgetManagerCLIView(bm_cliview_dc).initialize()
        # startup the view or not
        clview.cmdloop() if startup else None # Application CLI loop
        _ = "pause"
    except Exception as e:
        m = p3u.exc_err_msg(e)
        logger.error(m)
        raise
#endregion budman_app_cli_cmdloop() function
# ---------------------------------------------------------------------------- +
#region budman_app_start() function
def budman_app_start(testmode:bool=False):
    """Main function to run PyExcelBudget application."""
    try:
        # Here is where argv would be 
        # applied.
        # But for now, budman_app_cli is the only option to run.
        configure_logging(settings.app_name, logtest=testmode)
        logger.setLevel(logging.DEBUG)
        p3u.set_print_output(False)
        startup = not testmode
        budman_app_cli_cmdloop(startup=startup) # Application CLI loop
        _ = "pause"
    except Exception as e:
        m = p3u.exc_err_msg(e)
        logger.error(m)
        raise
#endregion budman_app_start() function
# ---------------------------------------------------------------------------- +
#region Local __main__ stand-alone
if __name__ == "__main__":
    try:
        budman_app_start() # Application Main()
    except Exception as e:
        m = p3u.exc_msg("__main__", e)
        logger.error(m)
    logger.info(f"Exiting {settings.app_name}...")
    exit(0)
#endregion Local __main__ stand-alone
# ---------------------------------------------------------------------------- +
#region tryit() function
def tryit() -> None:
    """Try something."""
    try:
        result = inspect.stack()[1][3]
        print(f"result: '{result}'")
    except Exception as e:
        print(f"Error: tryit(): {str(e)}")
#endregion tryit() function
# ---------------------------------------------------------------------------- +
